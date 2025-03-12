# -*- coding: utf-8 -*-
"""
some small functions which are useful when graphing things, 

Created on Thu Feb 11 18:01:48 2021

@author: dn1847
"""
import numpy as np

def get_csv_data_WK6500B(num_files='all', data_folder=None, exclude_subdirs=True):
    """
    Read in data from multiple .csv files and compile it into arrays.
    If data_folder == None, sets to current working directory.
    Uses specific names for files and column headers.
    Currently, files are numbered 1.csv, 2.csv, ..., 32.csv
    --> I need to update this to be a general file handler.
    
    Created on Wed Feb 20 15:11:49 2019
    
    @author: dn1847
    """
    import csv
    import os
    
    #file io
    data_folder = data_folder
    if data_folder==None:
        data_folder = os.getcwd()
    file_ext = '.csv'
    
    # get a list of the filenames
    fnames, fpaths = [], []
    if num_files == 'all':
        for root, dirs, files in os.walk(data_folder):
            #exclude subdirectories?
            if exclude_subdirs == True: #exclude all subdirectories
                dirs[:] = []
            elif type(exclude_subdirs) == list: #exclude specified subdirs
                dirs[:] = [dd for dd in dirs if dd not in exclude_subdirs]
            for f in files:
                if file_ext in f and 'tempData' not in f:
                    fnames.append(f)
                    fpaths.append(root)
                    print('adding file: ', f)
        num_files = len(fnames)
    elif type(num_files) == int:
        for i in range(1, num_files+1): #32 files
            fnames.append(str(i) + '.csv')
    # output files as dictionaries in list:
    all_data_dict_list = [] #list to store and return all the data

    #--------------------------------------------------------------------------
    #open all the csv data files and read data into dictionaries
    #--------------------------------------------------------------------------
    temperature_dict = None #will be populated if necessary, from external file
    for i in range(len(fnames)): #in range(1, num_files+1):#32 files
        ipath, ifile = fpaths[i], fnames[i]    
        #file_num = ifile
        if file_ext in ifile:
            fext = ''
        else:
            fext = file_ext
        full_file = os.path.join(ipath, ifile + fext)
        # print('full_file: ', full_file)
        try:
            with open(full_file, 'r', newline='') as f:
                reader = csv.reader(f)
                #read the header line
                header_line = next(reader)
                #make dict to store the column data from the csv file:
                data_dict = {'name' : ifile.split('.')[0]}
                for col_header in header_line:
                    data_dict[col_header] = []

                #check for a temperature column:
                #if there is not a temperature column, we need to add the temp
                #for each entry from the external <tempData> file.
                for hdr in header_line:
                    if 'temp' in hdr.casefold():
                        tempCol = True
                        print('temperature column exists in file: %s' %ifile)
                        break
                    else:
                        tempCol = False
                        
                if not tempCol:
                        print('temperature column does not exist in file: %s' %ifile)
                        data_dict['Average Temp (C)'] = [] #add a temperature mapping
                        if not temperature_dict:
                            #get the temps from external file
                            try:
                                temperature_dict = get_temperatures('tempData', data_folder, file_ext)
                                print('    --> importing temp data from \'/tempData.csv')
                            except UnboundLocalError as err:
                                print('***\nError retrieving temp data from \'tempData.csv\', \n***\n', err)
                                print('\n*** continue, assuming temp == 20 degC ***\n')
                                temperature_dict = {}
                                for ii in range(len(fnames)):
                                    temperature_dict[ii+1] = 20
                
                #now populate the lists in each field of the data_dict
                for line in reader:
                    for icol in range(len(line)):
                        ikey = header_line[icol]
                        data_dict[ikey].append(float(line[icol]))
                    if not tempCol: #add the temp from external file
                        #data_dict['Av.Temp'] = [] #initiate temperature mapping
                        data_dict['Average Temp (C)'].append(temperature_dict[fnames.index(ifile)+1])
#                        if not len(temperature_dict) == 0:
#                            data_dict['Average Temp (C)'].append(temperature_dict[fnames.index(ifile)+1])
#                        else:
#                            data_dict['Average Temp (C)'].append(None)

                all_data_dict_list.append(data_dict)
        
        except FileNotFoundError as err:
            print('File not found: ', err, '\nSkipping to next file...')
    
    return all_data_dict_list



def get_excel_data(filename='R_vs_f.xlsx', sheetNames='all', data_folder=None, compute_freq_averages=True):
    """
    Gets excel data, where the first row is a header and the data are in columns.
    Data is returned as a list of dictionaries.
    Each dictionary contains data from one sheet of the Excel workbook.
    Row headers in the original data file are used as the keys in each dict.
    
    INPUTS:
    filename: String of your filename describing the Excel workbook.
    unless in same folder as this .py execution.
    sheetNames: list of strings of the names of the worksheets required, within the
    Excel workbook described by filename.
    
    If data_folder == None, sets to current working directory.
    
    OUTPUTS:
    all_data_dict_list[]: A list of dictionaries from the Excel workbook. Each 
    dictionary holds the data from one worksheet of the Excel workbook. The keys to
    each dictionary are the column headers from the sheet.
    
    i.e.: all_data_dict_list = [[dict:sheet1], [dict:sheet2], ..., [dict:sheetn]]
    each dict is organised as:
        dictn = { name : 'sheet_name',
                  colA_header : [list of colA values],
                  colB_header : [list of colB values],
                  ...
                  colN_header : [list of colN values]
                 }        
    
    Created on Thurs Mar 28 2019
    
    @author: dn1847
    """    
    
#    import csv
#    import numpy as np
    from openpyxl import load_workbook

    if data_folder != None:
        fname = data_folder + '/' + filename
    else:
        fname = filename
    print('fname: ', fname)
#   #  Which column headers to read from:
#    col_headers = {'frequency_header' : 'freq1',
#                   'resistance_header' : 'Rs1',
#                   'impedance_header' : 'Z1',
#                   'angle_header' : 'phi1',
#                   'temp_header' : 'temp'
#                   }
#    
    #create container to store the data (one dictionary per worksheet)
    all_data_dict_list = []
    
    #open the file and get the row headers
    wb = load_workbook(filename=fname, data_only=True)
    if sheetNames == 'all':
        sheets = wb.sheetnames
    else:
        sheets = sheetNames
        
    for isheet in sheets:
        #make isheet active
        ws = wb[isheet]
        #make data storage
        wsdata = {}
        #get the sheet name
        wsdata['name'] = isheet
          
        #read data from worksheet 'ws':
        #make column iterator
        col_it = ws.columns
        #for each column read the column and append cell values to a list
        for icol in col_it:
            col_data = []
            for icell in icol[1:]:
                col_data.append(icell.value)
                
            #store the list in the dictionary mapped to the key from the 1st cell
            wsdata[icol[0].value.strip()] = col_data
        
        #write the worksheet data into the allData[] storage
        all_data_dict_list.append(wsdata)
    #close workbook
    wb.close()
    
    
    #--------------------------------------------------------------------------
    #for multiple readings at each frequency it is useful to average the data
    #--------------------------------------------------------------------------
    if compute_freq_averages == True:
        for dataset_dict in all_data_dict_list:
            #for each set of readings
            f_list = dataset_dict['freq1']
            #make a list of the starting indices of each group of readings
            new_freq_indices = [0]
            ii = 1
            while ii < len(f_list):
                if f_list[ii] == f_list[ii-1]:
                    #this entry is a repeated measurement at the same freq
                    pass
                elif f_list[ii] != f_list[ii-1]:
                    #this entry is at a new frequency
                    new_freq_indices.append(ii)
                ii+=1
            #now we have a list of the starting indices of each freq group
            #append extra index to end of list (necessary for averages)
            new_freq_indices.append(ii)
            
            #go through all the params and replace grouped data with averages
            keys = dataset_dict.keys()
            for k in keys:
                if k == 'name':
                    pass
                elif k in ['Date', 'Time']:
                    #choose the first entry, don't try to average
                    dataset_dict[k] = dataset_dict[k][0]
                else:
                    #print('k: ', k)
                    averages_list = []
                    tmp_list = dataset_dict[k] #a list of non-averaged data
                    #print(tmp_list[:10])
                    i0 = 0 #first index
                    for i in new_freq_indices[1:]:
                        #print('i0: %s, i: %s' %(i0, i))
                        #print('tmp_list[i0:i]: ',tmp_list[i0:i])
                        tmp_sum = sum(tmp_list[i0:i])
                        tmp_av = tmp_sum / (i - i0)
                        averages_list.append(tmp_av)
                        i0 = i
                    
                    #replace long list with averaged list
                    dataset_dict[k] = averages_list
    
    return all_data_dict_list

def get_temperatures(temperature_file = 'tempData', data_folder=None, file_ext='.csv'):
    import csv
    """Supplementary function: gets the <fileID> - <temperature> pairs from
    an external file <temperature_file> and returns them as a dict as:
        { int(fileID) : float(temperature) }
        
    File should be written STRICTLY in the following format and saved as
    a .csv in the same directory as the data files.
    
    'tempData.csv' MUST CONTAIN the following two columns (it may include
    more columns with non-conflicting column headers):
       | File Number | ... | Average Temp (C) | ... |
       |      1      | ... |     <temp1>      | ... |
       |      2      | ... |     <temp2>      | ... |
       |      ...    | ... |       ...        | ... |
       |      n      | ... |     <tempn>      | ... |
       
    'Average Temp (C)' should be the average temperature for a particular 
    execution of the experiment, with data stored in files corresponding to 
    the 'File Number' column.
    (File Number 1 corresponds to file '1.csv' in the same folder)
        
    """
    temperature_file_num_header = 'File Number'
    temperature_header = 'Average Temp (C)'
    
    #open the temperature file and read temps into a list with their file numbers
    temperature_file_name = temperature_file
    temperature_full_file = '%(folder)s\\%(file)s%(file_ext)s' % {
        'folder': data_folder, 'file': temperature_file_name, 'file_ext': file_ext }
    
    try:
        with open(temperature_full_file, newline='') as tf:
            t_reader = csv.reader(tf)
            #read the header line
            temp_header = next(t_reader)
            #identify the column indices for file number and average temperature
            file_num_col = temp_header.index(temperature_file_num_header)
            temp_col = temp_header.index(temperature_header)
            #fill a dict with {file number : average temp} entries
            temperature_dict = {}
            for row in t_reader:
                jfile = int(row[file_num_col])
                jtemp = float(row[temp_col])
                #temperature_list.append((ifile, itemp))
                temperature_dict[jfile] = jtemp
    except FileNotFoundError as err:
            print('***\nTemperature file not found! \nINCLUDE THE TEMPERATURE FILE \"tempData.csv\" IN THE SAME DIRECTORY AS THE DATA FILES\n: ', err)
            
    return temperature_dict

def peak_finder(data_dict_list, data_key, f_key=None):
    from scipy.signal import find_peaks
    '''
    Quick script to find the peaks in the data stored in data_dict_list, and
    returns a dict of lists (note the different structure to <data_dict_list>).

    Parameters
    ----------
    data_dict_list : list of dicts
        As compiled in <graphing_impedancePeaks.py>, this is a master list of
        all the impedance data read. i.e. coil names, frequencies, impedances...
    data_key : string
        a key from data_dict_list[i]. Peaks will be found in this dataset.
    f_key : string
        the data_key assigned to frequencies in data_dict_list. Returns empty lists
        if not assigned.
    Returns
    -------
    dpeaks : {[],[],[]}
        A dict of lists:
        {'name': [all the 'name' fields in data_dict_list],
         'data_peaks': [[peaks in <data_key> for item0],["" item1],...,["" itemn]]
         'f_peaks': [frequencies corresponding to the peaks]}
    '''
    ddata = data_dict_list
    dpeaks = {'name' : [], 'data_peaks' : [], 'f_peaks' : []}
    for item in ddata:
        zs = item[data_key]
        pks = find_peaks(zs)[0]
        zpks = [zs[pk] for pk in pks]
        dpeaks['name'].append(item['name'])
        dpeaks['data_peaks'].append(zpks)
        if f_key:
            fs = item[f_key]
            fpks = [fs[pk] for pk in pks]
            dpeaks['f_peaks'].append(fpks)
            
    return dpeaks

def peak_finder_dict_list(data_dict_list, data_key, f_key=None):
    from scipy.signal import find_peaks
    '''
    Finds the peaks in the data stored in data_dict_list, and returns a list
    of dicts in the same structure as data_dict_list

    Parameters
    ----------
    data_dict_list : list of dicts
        As compiled in <graphing_impedancePeaks.py>, this is a master list of
        all the impedance data read. i.e. coil names, frequencies, impedances...
    data_key : string
        a key from data_dict_list[i]. Peaks will be found in this dataset.
    f_key : string
        the data_key assigned to frequencies in data_dict_list. Returns empty lists
        if not assigned.
    Returns
    -------
    list of dicts. {'name': [all the 'name' fields in data_dict_list],
                    'data_peaks': [[peaks in <data_key> for item0],["" item1],...,["" itemn]]
                    'f_peaks': [frequencies corresponding to the peaks]}
    '''
    ddata = data_dict_list
    dpeaks = []
    idx = -1
    for item in ddata:
        idx += 1
        try:
            item_name = item['name']
        except KeyError:
            item_name = idx
            
        zs = item[data_key]
        pks = find_peaks(zs)[0]
        zpks = [zs[pk] for pk in pks]
        dpeaks.append({
            'name' : item_name,
            'data_peaks' : zpks,
            'peaks_of_data' : data_key,
            })
        if f_key:
            fs = item[f_key]
            fpks = [fs[pk] for pk in pks]
            dpeaks[idx]['f_peaks'] = fpks
    
    return dpeaks

def gradient_finder_dict_list(data_dict_list, data_key, f_key=None, end_search_at_peak=False):
    '''
    Finds the gradients in the data stored in data_dict_list, and returns 
    a list of dicts in the same structure as data_dict_list

    Parameters
    ----------
    data_dict_list : list of dicts
        As compiled in <graphing_impedancePeaks.py>, this is a master list of
        all the impedance data read. i.e. coil names, frequencies, impedances...
    data_key : string
        a key from data_dict_list[i]. Gradients will be found in this dataset.
    f_key : string
        the data_key assigned to frequencies in data_dict_list. 
        Sets spacing of the data items to 1 if omitted.
        Sets spacing of the data items to f_key if a single numeric value.
        i.e. np.gradient(data) will return (dn+1 - dn)/1 along data 
        instead of (dn+1 - dn)/(fn+1 - fn).
    Returns
    -------
    list of dicts. [{'name': [all the 'name' fields in data_dict_list],
                    'gradients': [[gradients along <data_key> for item0],["" item1],...,["" itemn]],
                    'max_grad': [[max gradient in <data_key> for item0],["" item1],...,["" itemn]]}]
    '''
    from scipy.signal import find_peaks
    
    ddata = data_dict_list
    dgrads = []
    
    idx = -1
    for item in ddata:
        idx += 1
        try:
            item_name = item['name']
        except KeyError:
            item_name = idx
        
        zs = item[data_key]
        
        if f_key is not None:
            if isinstance(f_key, (int,float)):
                xs = f_key
            elif isinstance(f_key, str):
                xs = item[f_key]
        else:
            xs = 1
        
        if end_search_at_peak:
            #index_pk = zs.index(max(zs))
            index_pk = find_peaks(zs)[0][0]
            grads = np.gradient(zs[:index_pk+1], xs[:index_pk+1])
        else:
            grads = np.gradient(zs, xs)

        maxgrad = max(grads)
        dgrads.append({
            'name' : item_name,
            'gradients' : grads,
            'max_grad' : maxgrad,
            })
    
    return dgrads
        
def pop_furthest_from_mean(popfrom):
    """
    calculates the mean of all elements (must be numeric types) and pops the 
    element furthest from it

    Parameters
    ----------
    popfrom : list
        list of numeric entries

    Returns
    -------
    newlist : list
        the original list <popfrom> with the popped item removed
    popindex : int
        the index of the popped item before removal
    popval : float
        the popped item

    """
    newlist = popfrom.copy()
    mn = np.mean(newlist)
    dist_from_mean = [np.abs(i-mn) for i in newlist]
    popindex = np.argmax(dist_from_mean)
    popval = newlist.pop(popindex)
    
    return newlist, popindex, popval

def read_in_data(data_folder, file_type='csv', exclude_subdirs=True, excel_fname=None, excel_sheets='all', convert_Z=False, compute_freq_averages=False):
    """read in data from multiple csv files in the same <data_folder>, or a 
    single excel file called <excel_fname>.
    
    -----------
    Parameters:
        data_folder
        type: string
            path to the folder containing the data files
        file_type
        type: string
            'csv' or 'excel'. If csv, the folder should contain csv files with
            an individual dataset in each. If excel, the excel file should 
            contain all the relevant datasets. The remaining kwargs will
            determine which data are read.
        exclude_subdirs
        type: bool or list
            passed to function 'get_csv_data_WK6500B()'.
            applies to .csv files only (because this uses all csv files in the 
            directory). Specify <True> to exclude all subdirectories below
            <data_folder>. False to include all subdirectories, or provide a 
            list of strings of directory names to exclude. 
        excel_fname
        type: string
            name of the excel file
        excel_sheets
            type: [string]
            list of strings of sheetnames to use, or 'all'
        convert_Z
        type: string or None
            Convert Z and phi data to resistance 'R' or reactance 'X'.
            Assumes there is impedance (Z) and phase angle (phi) data available
        compute_freq_averages
            type: bool
            if multiple data at the same frequency, use their average value
            
    -----------
    Returns:
        data_dict_list
        type: [dict]
            a list of dictionaries containing the data
    """    
    if file_type == 'csv':
        data_dict_list = get_csv_data_WK6500B(num_files='all', data_folder=data_folder, exclude_subdirs=exclude_subdirs)
        
    elif file_type == 'excel':
        data_dict_list = get_excel_data(filename=excel_fname, 
                                        sheetNames=excel_sheets, 
                                        data_folder=data_folder, 
                                        compute_freq_averages=compute_freq_averages)
    #all the data is now in data_dict_list. Keys are column headers, vals are 
    #lists of data
    
    #-----------------------------------------------------------------------------
    # some post import formatting anyone?
    #-----------------------------------------------------------------------------
    
    #convert impedance and angle data to AC resistance or reactance
    if convert_Z not in ['', False, 'no', None]:
        z_data = {}
        phi_data = {}
        print('\n*****\nLooking for IMPEDANCE and ANGLE data to convert to RESISTANCE...\n')
        for iDataSet in range(len(data_dict_list)):
            print('iDataSet: ', iDataSet, ' out of ', range(len(data_dict_list)))
            z_data[iDataSet] = data_dict_list[iDataSet]['Meas. Impedance (O)']
            phi_data[iDataSet] = data_dict_list[iDataSet]['Meas. Angle (°)']
            print('... found z_data and phi_data\n')
            
        if convert_Z in ['R', 'r', 'resistance', 'Resistance']:
            print('Converting Z and PHI to R...')
            r_data = {}
            for iZ in range(len(z_data)):   
                r_data[iZ] = z_data[iZ]*np.cos(np.radians(phi_data[iZ]))
                data_dict_list[iZ]['Calc. Resistance (O)'] = r_data[iZ]
        elif convert_Z in ['X', 'x', 'reactance', 'Reactance']:
            print('Converting Z and PHI to X...')
            reac_data = {}
            for iZ in range(len(z_data)):
                reac_data[iZ] = z_data[iZ]*np.sin(np.radians(phi_data[iZ]))
                data_dict_list[iZ]['Calc. Reactance (O)'] = reac_data[iZ]
                
    return data_dict_list


def short_name(long_name, tag, omit_tag=False, reverse_trim=False, keep_trailing_num=True):
     '''
    creates a shorter name from a longer one, based on a tag found in 
    the long_name. useful for legend entries.
 
     Parameters
     ----------
     long_name : str
         original long string to be shortened.
     tag : str
         a string of characters to look for in the long_name. Characters after
         tag will be omitted
     omit_tag : boolean, optional
         also omit the tag string from the short name.
         The default is False.
     reverse_trim : boolean, optional
         trim from the start of long_name to tag (reverse direction)
         The default is False.
     keep_trailing_num : boolean, optional
         If long_name has a trailing number it will be prefixed to short_name
         The default is True.
 
     Returns
     -------
     short_name : str
         short version of long_name.
 
     '''
     try:
         tag_index = long_name.index(tag)
         if omit_tag:
             if not reverse_trim:
                 short_name = long_name[:tag_index]
             elif reverse_trim:
                 short_name = long_name[tag_index+len(tag):]
         else: #i.e. omit_tag == False
             if not reverse_trim:
                 short_name = long_name[:tag_index+len(tag)]
             elif reverse_trim:
                 short_name = long_name[tag_index:]
     except ValueError as err:
         short_name = long_name
         print('Error when trying to shorten tag. Original tag will be used.')
         print('caught: ValueError: ', err)
     
     if keep_trailing_num:    
         if long_name[-1].isdigit(): #there's a number at the end of the string
             for s in np.arange(len(long_name)-1, -1, -1):
                 if not long_name[s].isdigit():
                     tail_num = long_name[s+1:] #the number at the end of the string
                     short_name = tail_num + '_' + short_name
                     break
     return short_name


def trim_name(long_name, tags, omit_tag=[False, False], keep_trailing_num=True):
    '''
    creates a shorter name from a longer one, by trimming characters at the
    start and end of the name, based on tags found in the long_name. useful 
    for legend entries.

    Parameters
    ----------
    long_name : str
        original long string to be shortened.
    tags : [str, str]
        list (len == 2) of character strings to look for at the start and end
        of the long_name. Characters before tags[0] and after tags[1] will be 
        cut. Use an empty string to cut zero characters.
    omit_tags : [boolean, boolean], optional
        also omit the tag(s) strings from the short name.
        The default is [False, False].
    keep_trailing_num : boolean, optional
        If long_name has a trailing number it will be prefixed to short_name
        The default is True.
    
    Returns
    -------
    short_name : str
        short version of long_name.
    '''
    tag_start = tags[0]
    tag_end = tags[1]

    # trim the back
    if(tag_end not in ['', None, False]):
       try:
           tag_index = long_name.index(tag_end)
           if omit_tag[1]:
               short_name = long_name[:tag_index]
           else: #i.e. omit_tag[1] == False
               short_name = long_name[:tag_index+len(tag_end)]
       except ValueError as err:
           short_name = long_name
           print('Error when trying to shorten tag end. End will not be trimmed.')
           print('caught: ValueError: ', err)
    
    # trim the front
    if(tag_start not in ['', None, False]):
        try:
            tag_index = long_name.index(tag_start)
            if omit_tag[0]:
                short_name = short_name[tag_index+len(tag_start):]
            else: #i.e. omit_tag[0] == False
                short_name = short_name[tag_index:]
        except ValueError as err:
            #short_name = long_name
            print('Error when trying to shorten tag. Start will not be trimmed.')
            print('caught: ValueError: ', err)

     
    if keep_trailing_num:    
        if long_name[-1].isdigit(): #there's a number at the end of the string
            for s in np.arange(len(long_name)-1, -1, -1):
                if not long_name[s].isdigit():
                    tail_num = long_name[s+1:] #the number at the end of the string
                    short_name = tail_num + '_' + short_name
                    break
    return short_name
 
